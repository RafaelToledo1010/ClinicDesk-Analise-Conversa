"""
analisar.py
Etapa 2 — analisa conversas com Claude API e gera Excel final
Variáveis de ambiente: ANTHROPIC_API_KEY
Modelo: claude-haiku-4-5-20251001
"""

import json
# VERSAO: HAIKU - 2026-05-25
import re
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

import anthropic
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# SYSTEM PROMPT GENÉRICO
# ─────────────────────────────────────────────

SYSTEM_PROMPT_BASE = """Você é um analista especializado em conversas de atendimento de clínicas médicas.

Você receberá conversas de WhatsApp ou Instagram entre uma clínica e seus contatos.
Sua tarefa é classificar cada conversa e retornar APENAS um JSON válido, sem markdown.

━━━ REGRAS DE CLASSIFICAÇÃO ━━━

[TIPO DE CONVERSA — escolha UM]
- paciente_externo: pessoa buscando consulta, procedimento ou exame
- lembrete_consulta: clínica enviou confirmação de horário já existente (sem nova negociação)
- cancelamento_reagendamento: paciente cancelou ou pediu para remarcar consulta já existente
- duvida_administrativa: reembolso, comprovante, documento, resultado de exame
- interno_fornecedor: empresa, fornecedor, contato comercial, sem contexto de paciente
- instagram_sem_interesse: mensagem via Instagram sem pedido de consulta
- conversa_fragmentada: mensagem isolada sem contexto suficiente para classificar

[AGENDAMENTO — critério estrito]
gerou_agendamento = true SOMENTE SE:
  A clínica confirmou data E horário explicitamente (ex: "ficou agendado para 15/02 às 10h")
  E o paciente confirmou ou não contestou.

gerou_agendamento = false SE:
  - Clínica ofereceu horário mas paciente NÃO respondeu
  - Paciente disse "vou pensar", "vou confirmar", "depois", "vou ver"
  - Conversa terminou sem confirmação dos dois lados
  - Era lembrete, reagendamento ou cancelamento
  - Paciente desistiu após saber informações (preço, convênio, etc)

gerou_agendamento = null SE:
  - Tipo não é paciente_externo

[CONVÊNIO — atenção especial]
Se paciente perguntou sobre convênio → clínica explicou que não aceita → paciente aceitou e agendou = true
Convênio só é motivo de NÃO agendamento se paciente DESISTIU após saber.

[MÉDICO]
Identifique o médico mencionado explicitamente na conversa.
Se não houver menção explícita → "nao_identificado".
Não infira pelo procedimento.

[PROCEDIMENTO]
Identifique o procedimento ou motivo do contato com base no que o paciente disse.
Seja preciso. Se não houver informação → "nao_especificado".

[MOTIVO DE NÃO AGENDAMENTO — escolha UM]
- convenio_nao_atendido: paciente queria convênio não aceito e desistiu
- financeiro: achou caro, sem condições de pagar
- distancia_localizacao: mora longe ou achou distante demais
- sem_retorno_paciente: clínica tentou contato e paciente não respondeu
- sem_retorno_clinica: paciente aguardou e clínica não retornou
- paciente_desistiu: disse que vai pensar/confirmar e sumiu
- outra_cidade: mora em outra cidade e não pode vir
- especialidade_errada: buscava especialidade que a clínica não oferece
- informacao_apenas: só queria tirar dúvida, sem intenção de agendar
- conversa_inconclusiva: contexto insuficiente para identificar motivo
- nao_aplicavel: tipo não é paciente_externo sem agendamento

[RESPONSÁVEL pelo não agendamento]
- clinica: clínica demorou, não retornou ou cometeu erro
- paciente: paciente desistiu, sumiu ou não respondeu
- externo: convênio, distância, especialidade, outra cidade
- nao_aplicavel: não se aplica

━━━ FORMATO DE SAÍDA — JSON APENAS, SEM MARKDOWN ━━━

{
  "tipo_conversa": "...",
  "medico": "nome do médico ou nao_identificado",
  "procedimento": "...",
  "gerou_agendamento": true | false | null,
  "motivo_nao_agendamento": "...",
  "responsavel": "clinica | paciente | externo | nao_aplicavel",
  "resumo": "1 frase objetiva descrevendo o que aconteceu",
  "evidencia": "trecho exato da conversa que justifica a classificação de agendamento"
}"""


# ─────────────────────────────────────────────
# APRENDIZADO ACUMULATIVO
# ─────────────────────────────────────────────

def carregar_aprendizado(caminho='aprendizado.json'):
    if not Path(caminho).exists():
        return []
    with open(caminho, encoding='utf-8') as f:
        return json.load(f)


def montar_prompt(aprendizado):
    if not aprendizado:
        return SYSTEM_PROMPT_BASE
    exemplos = aprendizado[-20:]
    bloco = '\n\n━━━ EXEMPLOS REAIS DE ERROS ANTERIORES — EVITE REPETIR ━━━\n'
    for i, ex in enumerate(exemplos, 1):
        bloco += f'\nExemplo {i}:\n'
        bloco += f'  Trecho: {ex.get("conversa_trecho", "")[:200]}\n'
        bloco += f'  Errado: {ex.get("classificacao_errada", "")}\n'
        bloco += f'  Correto: {ex.get("classificacao_correta", "")}\n'
    return SYSTEM_PROMPT_BASE + bloco


# ─────────────────────────────────────────────
# ANÁLISE DE UMA CONVERSA
# ─────────────────────────────────────────────

def analisar_conversa(client, conv, system_prompt, tentativas=3):
    texto = conv['conversa_texto']
    if len(texto) > 7000:
        texto = texto[:7000] + '\n[... truncado ...]'

    mensagem = (
        f"Paciente: {conv['paciente_nome']}\n"
        f"Telefone: {conv['paciente_tel']}\n"
        f"Plataforma: {conv['plataforma']}\n"
        f"Total de mensagens: {conv['total_mensagens']}\n\n"
        f"Conversa:\n{texto}"
    )

    for tentativa in range(1, tentativas + 1):
        try:
            resposta = client.messages.create(
                model='claude-haiku-4-5-20251001',
                max_tokens=1000,
                system=system_prompt,
                messages=[{'role': 'user', 'content': mensagem}]
            )
            txt = resposta.content[0].text.strip()
            txt = re.sub(r'```json\s*|\s*```', '', txt).strip()
            return json.loads(txt)

        except json.JSONDecodeError:
            if tentativa == tentativas:
                return _erro('JSON inválido')
            time.sleep(2)

        except anthropic.RateLimitError:
            time.sleep(60)

        except Exception as e:
            if tentativa == tentativas:
                return _erro(str(e)[:100])
            time.sleep(5)


def _erro(msg):
    return {
        'tipo_conversa': 'erro_processamento',
        'medico': None, 'procedimento': None,
        'gerou_agendamento': None, 'motivo_nao_agendamento': None,
        'responsavel': None, 'resumo': f'ERRO: {msg}', 'evidencia': None
    }


# ─────────────────────────────────────────────
# ANÁLISE DE TODAS AS CONVERSAS
# ─────────────────────────────────────────────

def analisar_todas(conversas, api_key, on_progresso=None):
    aprendizado = carregar_aprendizado()
    system_prompt = montar_prompt(aprendizado)
    client = anthropic.Anthropic(api_key=api_key)

    pendentes = [c for c in conversas if not c.get('analisado', False)]

    for i, conv in enumerate(pendentes):
        resultado = analisar_conversa(client, conv, system_prompt)
        conv.update(resultado)
        conv['analisado'] = True

        if on_progresso:
            on_progresso(i + 1, resultado)

        time.sleep(0.5)

    return conversas


# ─────────────────────────────────────────────
# GERAÇÃO DO EXCEL
# ─────────────────────────────────────────────

HDR_FILL  = PatternFill('solid', start_color='1F4E79')
HDR_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
ALT_FILL  = PatternFill('solid', start_color='EBF3FB')
WHT_FILL  = PatternFill('solid', start_color='FFFFFF')
GRN_FILL  = PatternFill('solid', start_color='C6EFCE')
RED_FILL  = PatternFill('solid', start_color='FFC7CE')
GRY_FILL  = PatternFill('solid', start_color='D9D9D9')
ORG_FILL  = PatternFill('solid', start_color='FCE4D6')
BLU_FILL  = PatternFill('solid', start_color='2E75B6')
YLW_FILL  = PatternFill('solid', start_color='FFEB9C')


def _hdr(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _body(ws, r, ncols, fill=None):
    f = fill or (ALT_FILL if r % 2 == 0 else WHT_FILL)
    for c in range(1, ncols + 1):
        cell = ws.cell(r, c)
        cell.fill = f
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.font = Font(name='Arial', size=9)


def _widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[col].width = w


def _secao(ws, row, texto, ncols=6):
    ws.cell(row, 1).value = texto
    ws.cell(row, 1).font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    ws.cell(row, 1).fill = BLU_FILL
    ws.cell(row, 1).alignment = Alignment(vertical='center')
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    ws.row_dimensions[row].height = 22


def _kpi(ws, row, label, valor, fill=None):
    ws.cell(row, 1).value = label
    ws.cell(row, 1).font = Font(bold=True, name='Arial', size=10)
    ws.cell(row, 2).value = valor
    ws.cell(row, 2).font = Font(bold=True, name='Arial', size=11)
    ws.cell(row, 2).alignment = Alignment(horizontal='center')
    if fill:
        ws.cell(row, 2).fill = fill


def gerar_excel(conversas, caminho_saida):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = 'conversas_completas'
    cols1 = ['Paciente', 'Telefone', 'Plataforma', 'Data Início', 'Data Fim',
             'Msgs', 'Tipo Conversa', 'Médico', 'Procedimento', 'Agendou?',
             'Responsável', 'Motivo Não Agendamento', 'Resumo', 'Evidência', 'Conversa']
    ws1.append(cols1)
    _hdr(ws1, 1, len(cols1))

    for conv in sorted(conversas, key=lambda x: x.get('data_inicio', '')):
        ag = conv.get('gerou_agendamento')
        ws1.append([
            conv.get('paciente_nome', ''),
            conv.get('paciente_tel', ''),
            conv.get('plataforma', ''),
            conv.get('data_inicio', ''),
            conv.get('data_fim', ''),
            conv.get('total_mensagens', 0),
            conv.get('tipo_conversa', ''),
            conv.get('medico', '') or '',
            conv.get('procedimento', '') or '',
            'Sim ✅' if ag is True else ('Não ❌' if ag is False else 'N/A'),
            conv.get('responsavel', '') or '',
            conv.get('motivo_nao_agendamento', '') or '',
            conv.get('resumo', '') or '',
            conv.get('evidencia', '') or '',
            conv.get('conversa_texto', '')[:2000],
        ])

    for r in range(2, len(conversas) + 2):
        ag = ws1.cell(r, 10).value
        tipo = str(ws1.cell(r, 7).value or '').lower()
        if ag == 'Sim ✅': fill = GRN_FILL
        elif ag == 'Não ❌': fill = RED_FILL
        elif any(t in tipo for t in ['lembrete', 'interno', 'instagram']): fill = GRY_FILL
        elif 'cancelamento' in tipo: fill = ORG_FILL
        else: fill = ALT_FILL if r % 2 == 0 else WHT_FILL
        _body(ws1, r, len(cols1), fill)

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f'A1:{get_column_letter(len(cols1))}1'
    _widths(ws1, {'A': 24, 'B': 16, 'C': 11, 'D': 12, 'E': 12, 'F': 6,
                  'G': 22, 'H': 20, 'I': 22, 'J': 10, 'K': 12,
                  'L': 38, 'M': 50, 'N': 50, 'O': 60})

    ws2 = wb.create_sheet('analise_executiva')

    pacientes    = [c for c in conversas if c.get('tipo_conversa') == 'paciente_externo']
    lembretes    = [c for c in conversas if c.get('tipo_conversa') == 'lembrete_consulta']
    cancelam     = [c for c in conversas if c.get('tipo_conversa') == 'cancelamento_reagendamento']
    internos     = [c for c in conversas if c.get('tipo_conversa') in ['interno_fornecedor', 'instagram_sem_interesse']]
    duvidas      = [c for c in conversas if c.get('tipo_conversa') == 'duvida_administrativa']
    fragmentados = [c for c in conversas if c.get('tipo_conversa') == 'conversa_fragmentada']
    ag_list      = [c for c in pacientes if c.get('gerou_agendamento') is True]
    nag_list     = [c for c in pacientes if c.get('gerou_agendamento') is False]
    total        = len(conversas)
    n_ag         = len(ag_list)
    n_nag        = len(nag_list)
    taxa         = f'{n_ag / len(pacientes) * 100:.1f}%' if pacientes else '0%'

    _secao(ws2, 1, '📊 ANÁLISE EXECUTIVA — CLINICDESK ANÁLISES')
    ws2.row_dimensions[1].height = 26

    _kpi(ws2, 3, 'Total de conversas no período', total)
    r = 4
    tipos_dist = [
        ('  Pacientes',                           len(pacientes),    None),
        ('  Lembretes de consulta',               len(lembretes),    GRY_FILL),
        ('  Internos / fornecedores / Instagram', len(internos),     GRY_FILL),
        ('  Cancelamentos / reagendamentos',      len(cancelam),     ORG_FILL),
        ('  Dúvidas administrativas',             len(duvidas),      GRY_FILL),
        ('  Conversas fragmentadas',              len(fragmentados), GRY_FILL),
    ]
    soma = 0
    for label, valor, fill in tipos_dist:
        _kpi(ws2, r, label, valor, fill)
        soma += valor
        r += 1

    ws2.cell(r, 1).value = '  TOTAL (conferência)'
    ws2.cell(r, 1).font = Font(bold=True, name='Arial', size=10)
    ws2.cell(r, 2).value = soma
    ws2.cell(r, 2).font = Font(bold=True, name='Arial', size=11)
    ws2.cell(r, 2).alignment = Alignment(horizontal='center')
    ws2.cell(r, 2).fill = YLW_FILL
    r += 2

    _secao(ws2, r, '🏥 DETALHAMENTO — PACIENTES', 4)
    r += 1
    _kpi(ws2, r, 'Pacientes', len(pacientes))
    r += 1
    _kpi(ws2, r, '  ✅ Agendaram', n_ag, GRN_FILL)
    r += 1
    _kpi(ws2, r, '  ❌ Não agendaram', n_nag, RED_FILL)
    r += 1
    _kpi(ws2, r, '  Taxa de conversão', taxa, YLW_FILL)
    r += 2

    _secao(ws2, r, '🔍 MOTIVOS DE NÃO AGENDAMENTO', 4)
    r += 1
    for ci, lbl in enumerate(['Motivo', 'Qtd', '% do total não agendados'], 1):
        ws2.cell(r, ci).value = lbl
    _hdr(ws2, r, 3)
    r += 1
    motivos = Counter(c.get('motivo_nao_agendamento') for c in nag_list
                      if c.get('motivo_nao_agendamento') not in [None, '', 'nao_aplicavel'])
    for motivo, cnt in motivos.most_common():
        ws2.cell(r, 1).value = motivo
        ws2.cell(r, 2).value = cnt
        ws2.cell(r, 3).value = f'{cnt / n_nag * 100:.1f}%' if n_nag else '0%'
        _body(ws2, r, 3)
        r += 1

    r += 1
    _secao(ws2, r, '👨‍⚕️ AGENDAMENTOS POR MÉDICO', 5)
    r += 1
    for ci, lbl in enumerate(['Médico', 'Agendamentos', 'Não agendamentos', 'Total', 'Taxa'], 1):
        ws2.cell(r, ci).value = lbl
    _hdr(ws2, r, 5)
    r += 1
    todos_medicos = sorted(set(c.get('medico') or 'nao_identificado' for c in pacientes))
    for med in todos_medicos:
        ag_m  = sum(1 for c in pacientes if (c.get('medico') or 'nao_identificado') == med and c.get('gerou_agendamento') is True)
        nag_m = sum(1 for c in pacientes if (c.get('medico') or 'nao_identificado') == med and c.get('gerou_agendamento') is False)
        tot_m = ag_m + nag_m
        ws2.cell(r, 1).value = 'Não identificado' if med == 'nao_identificado' else med
        ws2.cell(r, 2).value = ag_m
        ws2.cell(r, 3).value = nag_m
        ws2.cell(r, 4).value = tot_m
        ws2.cell(r, 5).value = f'{ag_m / tot_m * 100:.0f}%' if tot_m else '0%'
        _body(ws2, r, 5)
        r += 1

    ws2.cell(r, 1).value = 'TOTAL'
    ws2.cell(r, 2).value = n_ag
    ws2.cell(r, 3).value = n_nag
    ws2.cell(r, 4).value = len(pacientes)
    ws2.cell(r, 5).value = taxa
    for c in range(1, 6):
        ws2.cell(r, c).fill = YLW_FILL
        ws2.cell(r, c).font = Font(bold=True, name='Arial', size=9)
    r += 1

    r += 1
    falhas = [c for c in nag_list if c.get('responsavel') == 'clinica']
    _secao(ws2, r, f'🔴 FALHAS DA CLÍNICA — {len(falhas)} casos', 4)
    r += 1
    for ci, lbl in enumerate(['Paciente', 'Motivo', 'Data'], 1):
        ws2.cell(r, ci).value = lbl
    _hdr(ws2, r, 3)
    r += 1
    for c in falhas:
        ws2.cell(r, 1).value = c.get('paciente_nome', '')
        ws2.cell(r, 2).value = c.get('motivo_nao_agendamento', '')
        ws2.cell(r, 3).value = c.get('data_inicio', '')
        _body(ws2, r, 3, RED_FILL)
        r += 1

    _widths(ws2, {'A': 50, 'B': 16, 'C': 20, 'D': 10, 'E': 10})

    wb.save(caminho_saida)
    return caminho_saida
